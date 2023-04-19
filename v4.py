import socket
import openpyxl


def impressora(codigo_zpl):
    # Endereço IP da impressora
    HOST = '192.168.152.52'
    # Porta padrão de comunicação com a impressora
    PORT = 9100

    # Conecta-se à impressora
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.connect((HOST, PORT))

    # Envia o código ZPL para a impressora
    s.sendall(codigo_zpl.encode())

    # Fecha a conexão com a impressora
    s.close()


def codigo_impressora(concatena, formatado):
    codigo_zpl = 'CT~~CD,~CC^~CT~  \n' \
                 '^XA~TA000~JSN^LT0^MNW^MTT^PON^PMN^LH0,0^JMA^PR4,4~SD20^JUS^LRN^CI0^XZ \n' \
                 '^XA \n' \
                 '^MMT \n' \
                 '^PW799 \n' \
                 '^LL0400 \n' \
                 '^LS0 \n' \
                 '^BY3,3,117^FT633,100^BCI,,N,N\n' \
                 '^FD>:' + concatena + '^FS \n' \
                 '^FT736,297^A0I,87,93^FH\^FD' + formatado + '^FS \n' \
                 '^PQ1,0,1,Y^XZ \n'
    return codigo_zpl

aux = True

escolha = input("Digite 1 para Gerar em Massa \nDigite 2 para Reemprimir \nDigite 3 para Sair\n")

while aux:

    if escolha == '1':
        # workbook = openpyxl.load_workbook('/Users/Public/Documents/ENDEREÇOS.xlsx')
        workbook = openpyxl.load_workbook('Q:/#/Suporte TI/Fred/Automacao-logistica/test.xlsx')
        sheet = workbook.active

        # Inicia com a célula A2
        row_number = 2
        var = sheet.cell(row=row_number, column=1).value

        # Loop enquanto a célula atual não estiver vazia
        while var is not None:
            concatena = var + '5'
            formatado = "-".join([var[:2], var[2:4], var[4], var[5:]])

            # Layout da etiqueta
            codigo_zpl = codigo_impressora(concatena, formatado)

            impressora(codigo_zpl)

            # Atualiza para a próxima célula da coluna A
            row_number += 1
            var = sheet.cell(row=row_number, column=1).value
            aux = False


    elif escolha == '2':
        var = input("Insira o numero para Reemprimir : ")
        concatena = var + '5'
        formatado = "-".join([var[:2], var[2:4], var[4], var[5:]])

        # Layout da etiqueta
        codigo_zpl = codigo_impressora(concatena, formatado)

        impressora(codigo_zpl)
        aux = False

    elif escolha == '3':
        aux = False
    else :
        print("Insira um digito valido! ")
        escolha = input()
