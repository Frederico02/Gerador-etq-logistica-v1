import socket
import openpyxl

def busca_linha(inicio, planilha):
    while True:
        # Procura a linha onde a variável "inicio" está
        row_number = 0
        for row in planilha.iter_rows():
            row_number += 1
            if row[0].value == inicio:
                return row_number + 1
        else:
            print("Valor não encontrado.")
            novo_inicio = input("Tente novamente : ")
            print(f"Buscando por '{novo_inicio}'...")
            return busca_linha(novo_inicio, planilha)


def impressora(codigo_zpl):

    # Configuração da porta da impressora
    port = "LPT1"  # Altere para a porta correta da sua impressora
    with open(port, "wb") as printer:
        printer.write(codigo_zpl.encode("utf-8"))



def codigo_impressora(concatena, formatado):
    codigo_zpl = 'CT~~CD,~CC^~CT~  \n' \
                 '^XA~TA000~JSN^LT0^MNW^MTT^PON^PMN^LH0,0^JMA^PR4,4~SD20^JUS^LRN^CI0^XZ \n' \
                 '^XA \n' \
                 '^MMT \n' \
                 '^PW799 \n' \
                 '^LL0400 \n' \
                 '^LS0 \n' \
                 '^BY3,3,117^FT633,100^BCI,,N,N\n' \
                 '^FD>:' + concatena + '^FS \n' \
                 '^FT736,297^A0I,87,93^FH\^FD' + formatado + '^FS \n' \
                 '^PQ1,0,1,Y^XZ \n'
    return codigo_zpl


while True:
    escolha = input("Digite 1 para Gerar em Massa \nDigite 2 para Reemprimir \nDigite 3 para selecionar um intervalo \nDigite 4 para Sair\n\n")

    workbook = openpyxl.load_workbook('Q:/#/Suporte TI/Fred/Automacao-logistica/ENDEREÇOS.xlsx')
    sheet = workbook.active

    if escolha == '1':

        # Inicia com a célula A2
        row_number = 2
        var = sheet.cell(row=row_number, column=1).value

        # Loop enquanto a célula atual não estiver vazia
        while var is not None:
            concatena = '05' + var
            formatado = "-".join([var[:2], var[2:5], var[5], var[6:]])

            # Layout da etiqueta
            codigo_zpl = codigo_impressora(concatena, formatado)

            impressora(codigo_zpl)

            # Atualiza para a próxima célula da coluna A
            row_number += 1
            var = sheet.cell(row=row_number, column=1).value


    elif escolha == '2':
        var = input("Insira o numero para Reemprimir : ")
        concatena = '05' + var
        formatado = "-".join([var[:2], var[2:5], var[5], var[6:]])

        # Layout da etiqueta
        codigo_zpl = codigo_impressora(concatena, formatado)

        impressora(codigo_zpl)
    elif escolha == '3':
        inicio = input("Insira a posição que deseja iniciar : ")
        posInicial = (busca_linha(inicio, sheet))

        fim = input("Insira a posição que deseja Parar : ")
        porFinal = (busca_linha(fim, sheet))


        # Inicia com a célula A2
        row_number = posInicial
        var = sheet.cell(row=row_number, column=1).value
        i = posInicial
        # Loop enquanto a célula atual não estiver vazia
        while i <= porFinal:
            concatena = '05' + var
            formatado = "-".join([var[:2], var[2:5], var[5], var[6:]])

            # Layout da etiqueta
            codigo_zpl = codigo_impressora(concatena, formatado)

            impressora(codigo_zpl)

            # Atualiza para a próxima célula da coluna A
            row_number += 1
            var = sheet.cell(row=row_number, column=1).value

            i+=1

    elif escolha == '4':
        break
    else :
        print("Insira um digito valido! ")
        escolha = input()
